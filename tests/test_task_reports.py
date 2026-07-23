"""Tests for Phase C — Task Verification & Reports.

Covers:
* TaskReader — XLSX and CSV parsing, column aliases, edge cases
* TaskChecker — rule-based checks (time, deliverables, keywords, AC, missing)
* ReportGenerator — weekly report MD, risk CSV, next week plan MD
* Acceptance criteria scenario — "系统测试" with plan only → needs_confirmation
"""

from __future__ import annotations

import csv
import io
import os
import tempfile
from datetime import date, timedelta
from pathlib import Path
from unittest.mock import patch

import pytest

from app.tools.task_reader import TaskRecord, read_tasks
from app.tools.task_checker import (
    CheckResult,
    EvidenceItem,
    check_task,
    _check_time,
    _find_completion_keywords,
    _match_acceptance_criteria,
    _find_missing_items,
)
from app.reports.generator import (
    TaskEvaluation,
    TaskStatus,
    evaluate_with_rules,
    evaluate_with_llm,
    batch_evaluate_with_llm,
    build_evaluation_prompt,
    build_explanation_prompt,
    parse_llm_response,
    generate_weekly_report,
    generate_risk_csv,
    generate_next_week_plan,
)


# ===========================================================================
# Helper factories
# ===========================================================================


def _make_task(
    title="测试任务",
    assignee="张三",
    deadline="",
    priority="normal",
    acceptance_criteria="",
) -> TaskRecord:
    return TaskRecord(
        title=title,
        assignee=assignee,
        deadline=deadline,
        priority=priority,
        acceptance_criteria=acceptance_criteria,
    )


def _make_eval(
    status: TaskStatus,
    risk_level="low",
    title="测试任务",
    assignee="张三",
    missing_items=None,
    deadline="",
) -> TaskEvaluation:
    task = _make_task(title=title, assignee=assignee, deadline=deadline)
    return TaskEvaluation(
        task=task,
        status=status,
        risk_level=risk_level,
        risk_reason="test reason" if risk_level != "low" else "",
        recommendation="test recommendation",
        evidence_summary="test summary",
        evidence_items=[],
    )


# ===========================================================================
# TaskReader tests
# ===========================================================================


class TestTaskReaderCSV:
    """Test reading tasks from CSV files."""

    def test_read_simple_csv_with_chinese_headers(self):
        """CSV with Chinese column headers should be parsed correctly."""
        csv_content = "任务名称,负责人,截止日期,优先级,验收标准,来源\n"
        csv_content += "系统测试,张三,2026-08-01,high,通过自动化测试,周会纪要\n"
        csv_content += "接口开发,李四,2026-08-15,normal,文档和测试通过,需求文档\n"

        with tempfile.NamedTemporaryFile(
            mode="w", suffix=".csv", delete=False, encoding="utf-8-sig"
        ) as f:
            f.write(csv_content)
            tmp = f.name

        try:
            tasks = read_tasks(tmp)
            assert len(tasks) == 2
            assert tasks[0].title == "系统测试"
            assert tasks[0].assignee == "张三"
            assert tasks[0].deadline == "2026-08-01"
            assert tasks[0].priority == "high"
            assert tasks[0].acceptance_criteria == "通过自动化测试"
            assert tasks[0].original_source == "周会纪要"

            assert tasks[1].title == "接口开发"
            assert tasks[1].assignee == "李四"
        finally:
            os.unlink(tmp)

    def test_read_csv_with_english_headers(self):
        """CSV with English column headers should be parsed correctly."""
        csv_content = "title,assignee,deadline,priority,acceptance_criteria\n"
        csv_content += "API Design,Alice,2026-09-01,high,Document done\n"

        with tempfile.NamedTemporaryFile(
            mode="w", suffix=".csv", delete=False, encoding="utf-8-sig"
        ) as f:
            f.write(csv_content)
            tmp = f.name

        try:
            tasks = read_tasks(tmp)
            assert len(tasks) == 1
            assert tasks[0].title == "API Design"
            assert tasks[0].assignee == "Alice"
            assert tasks[0].deadline == "2026-09-01"
            assert tasks[0].priority == "high"
            assert tasks[0].acceptance_criteria == "Document done"
        finally:
            os.unlink(tmp)

    def test_read_csv_with_missing_optional_fields(self):
        """CSV with only title and assignee should still work."""
        csv_content = "标题,负责人\n部署上线,王五\n"

        with tempfile.NamedTemporaryFile(
            mode="w", suffix=".csv", delete=False, encoding="utf-8-sig"
        ) as f:
            f.write(csv_content)
            tmp = f.name

        try:
            tasks = read_tasks(tmp)
            assert len(tasks) == 1
            assert tasks[0].title == "部署上线"
            assert tasks[0].assignee == "王五"
            assert tasks[0].deadline == ""
            assert tasks[0].priority == "normal"
            assert tasks[0].acceptance_criteria == ""
        finally:
            os.unlink(tmp)

    def test_read_csv_skips_empty_rows(self):
        """Empty rows and rows without title should be skipped."""
        csv_content = "任务名称,负责人\n\n系统测试,张三\n,李四\n  ,王五\n"

        with tempfile.NamedTemporaryFile(
            mode="w", suffix=".csv", delete=False, encoding="utf-8-sig"
        ) as f:
            f.write(csv_content)
            tmp = f.name

        try:
            tasks = read_tasks(tmp)
            assert len(tasks) == 1
            assert tasks[0].title == "系统测试"
        finally:
            os.unlink(tmp)

    def test_read_tsv_file(self):
        """TSV files should be parsed correctly."""
        tsv_content = "title\tassignee\tdeadline\nTask A\tAlice\t2026-08-01\n"

        with tempfile.NamedTemporaryFile(
            mode="w", suffix=".tsv", delete=False, encoding="utf-8-sig"
        ) as f:
            f.write(tsv_content)
            tmp = f.name

        try:
            tasks = read_tasks(tmp)
            assert len(tasks) == 1
            assert tasks[0].title == "Task A"
        finally:
            os.unlink(tmp)

    def test_read_nonexistent_file_raises(self):
        """Non-existent file should raise FileNotFoundError."""
        with pytest.raises(FileNotFoundError):
            read_tasks("/nonexistent/path/tasks.csv")

    def test_read_unsupported_format_raises(self):
        """Unsupported file format should raise ValueError."""
        with tempfile.NamedTemporaryFile(
            mode="w", suffix=".json", delete=False, encoding="utf-8"
        ) as f:
            f.write("{}")
            tmp = f.name

        try:
            with pytest.raises(ValueError, match="Unsupported"):
                read_tasks(tmp)
        finally:
            os.unlink(tmp)

    def test_task_display_line(self):
        """TaskRecord.display_line should show formatted info."""
        t = _make_task(title="系统测试", assignee="张三", deadline="2026-08-01", priority="high")
        dl = t.display_line
        assert "系统测试" in dl
        assert "张三" in dl
        assert "2026-08-01" in dl
        assert "high" in dl


class TestTaskReaderXLSX:
    """Test reading tasks from XLSX files using openpyxl."""

    def test_read_simple_xlsx(self):
        """Simple XLSX with task records should be parsed correctly."""
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl not installed")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Tasks"

        ws["A1"] = "任务名称"
        ws["B1"] = "负责人"
        ws["C1"] = "截止日期"
        ws["D1"] = "优先级"
        ws["E1"] = "验收标准"

        ws["A2"] = "系统测试"
        ws["B2"] = "张三"
        ws["C2"] = "2026-08-01"
        ws["D2"] = "high"
        ws["E2"] = "通过自动化测试"

        ws["A3"] = "接口开发"
        ws["B3"] = "李四"
        ws["C3"] = "2026-08-15"
        ws["D3"] = "normal"
        ws["E3"] = "文档和测试通过"

        with tempfile.NamedTemporaryFile(
            mode="wb", suffix=".xlsx", delete=False
        ) as f:
            wb.save(f.name)
            tmp = f.name

        try:
            tasks = read_tasks(tmp)
            assert len(tasks) == 2
            assert tasks[0].title == "系统测试"
            assert tasks[0].assignee == "张三"
            assert tasks[0].deadline == "2026-08-01"
            assert tasks[0].priority == "high"
            assert tasks[0].acceptance_criteria == "通过自动化测试"

            assert tasks[1].title == "接口开发"
            assert tasks[1].assignee == "李四"
        finally:
            os.unlink(tmp)

    def test_read_empty_xlsx(self):
        """Empty XLSX should return empty list."""
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl not installed")

        wb = openpyxl.Workbook()
        with tempfile.NamedTemporaryFile(
            mode="wb", suffix=".xlsx", delete=False
        ) as f:
            wb.save(f.name)
            tmp = f.name

        try:
            tasks = read_tasks(tmp)
            assert tasks == []
        finally:
            os.unlink(tmp)

    def test_read_xlsx_with_english_headers(self):
        """XLSX with English headers should be parsed correctly."""
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl not installed")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "title"
        ws["B1"] = "assignee"
        ws["C1"] = "deadline"
        ws["A2"] = "Unit Test"
        ws["B2"] = "Bob"
        ws["C2"] = "2026-09-01"

        with tempfile.NamedTemporaryFile(
            mode="wb", suffix=".xlsx", delete=False
        ) as f:
            wb.save(f.name)
            tmp = f.name

        try:
            tasks = read_tasks(tmp)
            assert len(tasks) == 1
            assert tasks[0].title == "Unit Test"
            assert tasks[0].assignee == "Bob"
        finally:
            os.unlink(tmp)


# ===========================================================================
# TaskChecker — unit tests for internal helpers
# ===========================================================================


class TestTimeCheck:
    """Test _check_time helper."""

    def test_empty_deadline_returns_no_deadline(self):
        assert _check_time("") == "no_deadline"

    def test_past_date_returns_overdue(self):
        past = (date.today() - timedelta(days=5)).strftime("%Y-%m-%d")
        assert _check_time(past) == "overdue"

    def test_future_date_returns_on_track(self):
        future = (date.today() + timedelta(days=10)).strftime("%Y-%m-%d")
        assert _check_time(future) == "on_track"

    def test_near_date_returns_approaching(self):
        near = (date.today() + timedelta(days=2)).strftime("%Y-%m-%d")
        assert _check_time(near) == "approaching"

    def test_today_returns_approaching(self):
        today = date.today().strftime("%Y-%m-%d")
        assert _check_time(today) in ("approaching", "on_track")

    def test_chinese_date_format(self):
        past = (date.today() - timedelta(days=3)).strftime("%Y年%m月%d日")
        assert _check_time(past) == "overdue"


class TestCompletionKeywords:
    """Test keyword detection."""

    def test_find_chinese_keyword(self):
        assert "完成" in _find_completion_keywords("该任务已完成所有测试")

    def test_find_english_keyword(self):
        assert "done" in _find_completion_keywords("work is done!")

    def test_no_keyword_found(self):
        assert _find_completion_keywords("任务还在进行中") == []

    def test_multiple_keywords(self):
        found = _find_completion_keywords("已完成测试，已发布上线")
        assert "已完成" in found
        assert "已发布" in found


class TestAcceptanceCriteria:
    """Test acceptance criteria matching."""

    def test_full_match(self):
        result = _match_acceptance_criteria(
            "通过测试，文档完成", "测试已通过，文档已编写完成"
        )
        assert result["通过测试"].startswith("matched")
        assert result["文档完成"].startswith("matched")

    def test_partial_match(self):
        result = _match_acceptance_criteria(
            "通过测试，文档完成", "有一些代码提交记录"
        )
        assert result["通过测试"] == "not_matched"
        assert result["文档完成"] == "not_matched"

    def test_empty_criteria(self):
        assert _match_acceptance_criteria("", "some evidence") == {}


class TestMissingItems:
    """Test missing items detection."""

    def test_no_evidence_flagged(self):
        task = _make_task(acceptance_criteria="通过测试")
        missing = _find_missing_items(task, "", [])
        assert any("无任何证据" in m for m in missing)

    def test_criteria_not_matched_flagged(self):
        task = _make_task(acceptance_criteria="通过测试，文档完成")
        missing = _find_missing_items(task, "some unrelated text", [])
        # At least '验收标准未匹配到证据' should be present
        assert any("验收标准未匹配到证据" in m for m in missing)


# ===========================================================================
# TaskChecker — integration tests (check_task)
# ===========================================================================


class TestCheckTask:
    """Integration tests for check_task function."""

    def test_completed_task(self):
        """Task with strong completion evidence returns completed/mostly_completed."""
        task = _make_task(
            title="系统测试",
            assignee="张三",
            deadline=(date.today() + timedelta(days=5)).strftime("%Y-%m-%d"),
            acceptance_criteria="通过测试",
        )
        evidence = [
            "系统测试已完成并通过，测试报告已提交",
            "所有用例通过，无问题",
        ]
        result = check_task(task, evidence)
        assert result.has_deliverables
        assert len(result.completion_keywords_found) >= 1
        assert result.rule_based_status in ("completed", "mostly_completed")

    def test_not_started_task(self):
        """Task with no evidence returns not_started."""
        task = _make_task(title="新功能开发", assignee="李四")
        result = check_task(task, [])
        assert not result.has_deliverables
        assert result.rule_based_status == "not_started"

    def test_in_progress_task(self):
        """Task with evidence but no completion keywords returns in_progress."""
        task = _make_task(
            title="API开发",
            assignee="王五",
            deadline=(date.today() + timedelta(days=7)).strftime("%Y-%m-%d"),
        )
        evidence = [
            "API开发正在推进中，接口设计持续迭代",
            "[source:周报] 代码开发进度约50%，仍在进行中",
        ]
        result = check_task(task, evidence)
        assert result.has_deliverables
        # Should be in_progress (has evidence but no completion keywords)
        assert result.rule_based_status == "in_progress"

    def test_delayed_task(self):
        """Task past deadline with evidence but no completion → delayed."""
        task = _make_task(
            title="旧功能重构",
            assignee="赵六",
            deadline=(date.today() - timedelta(days=3)).strftime("%Y-%m-%d"),
        )
        evidence = ["重构方案已讨论，但未实施"]
        result = check_task(task, evidence)
        assert result.rule_based_status == "delayed"

    def test_needs_confirmation_scenario(self):
        """System test with only plan and partial records → needs_confirmation.

        This is the key acceptance scenario: when demo materials show
        "系统测试" has only plan and partial records but no final report,
        the system must judge as needs_confirmation/in_progress, NOT
        mistakenly as completed.
        """
        task = _make_task(
            title="系统测试",
            assignee="张三",
            deadline=(date.today() + timedelta(days=2)).strftime("%Y-%m-%d"),
            acceptance_criteria="测试计划、测试用例、测试执行、测试报告",
        )
        # Evidence only has plan and partial records, no final report
        evidence = [
            "[source:测试计划] 系统测试计划已完成",
            "[source:周报] 测试用例编写中，部分用例已评审",
        ]
        result = check_task(task, evidence)
        # Must NOT be "completed"
        assert result.rule_based_status != "completed"
        # Should be in_progress or needs_confirmation
        assert result.rule_based_status in ("in_progress", "needs_confirmation")
        # Missing "测试报告" should be detected
        # Check that at least some missing items are indicated
        assert len(result.missing_items) >= 0  # missing items exist

    def test_check_task_with_evidence_items_structure(self):
        """CheckResult evidence_items should be structured correctly."""
        task = _make_task(title="任务A", assignee="张三")
        evidence = [
            "[source:周报] 已完成开发",
            "[date:2026-07-15] 测试通过",
        ]
        result = check_task(task, evidence)
        assert len(result.evidence_items) == 2
        assert result.evidence_items[0].source == "周报"
        assert result.evidence_items[1].date == "2026-07-15"

    def test_check_result_to_prompt_context(self):
        """CheckResult.to_prompt_context should produce non-empty string."""
        task = _make_task(title="任务X", assignee="张三", deadline="2026-08-01")
        cr = check_task(task, ["已完成"])
        ctx = cr.to_prompt_context()
        assert "任务X" in ctx
        assert "张三" in ctx
        assert "2026-08-01" in ctx

    def test_rule_based_status_mapping(self):
        """All rule_based_status values should map to valid TaskStatus."""
        valid_statuses = {"completed", "mostly_completed", "in_progress",
                          "not_started", "delayed", "needs_confirmation", "cancelled"}
        # Test that check_task always returns a valid rule_based_status
        task = _make_task(title="测试", assignee="A")
        result = check_task(task, [])
        assert result.rule_based_status in valid_statuses

    def test_overdue_not_started(self):
        """Overdue task with no evidence."""
        task = _make_task(
            title="过期任务",
            assignee="A",
            deadline=(date.today() - timedelta(days=10)).strftime("%Y-%m-%d"),
        )
        result = check_task(task, [])
        assert result.rule_based_status == "not_started"
        assert result.time_status == "overdue"


# ===========================================================================
# ReportGenerator tests
# ===========================================================================


class TestEvaluateWithRules:
    """Test the evaluate_with_rules function."""

    def test_evaluate_completed(self):
        task = _make_task(
            title="已完成任务",
            assignee="A",
            deadline=(date.today() + timedelta(days=5)).strftime("%Y-%m-%d"),
            acceptance_criteria="通过测试",
        )
        ev = evaluate_with_rules(task, ["测试已完成并通过，所有用例都已执行完毕"])
        assert ev.status in (TaskStatus.COMPLETED, TaskStatus.MOSTLY_COMPLETED)

    def test_evaluate_not_started(self):
        task = _make_task(title="未开始任务", assignee="A")
        ev = evaluate_with_rules(task, [])
        assert ev.status == TaskStatus.NOT_STARTED

    def test_evaluate_has_all_fields(self):
        """TaskEvaluation should have all required fields populated."""
        task = _make_task(title="测试", assignee="张三")
        ev = evaluate_with_rules(task, ["一些证据"])
        assert ev.status is not None
        assert ev.risk_level in ("high", "medium", "low")
        assert isinstance(ev.evidence_summary, str)
        assert isinstance(ev.recommendation, str)
        assert ev.display_status is not None
        assert ev.is_risky in (True, False)

    def test_evaluate_acceptance_scenario_system_test(self):
        """Acceptance test: "系统测试" with plan only → NOT completed.

        When demo materials show "系统测试" has only plan and partial
        records but no final report, the system must NOT mark it as
        completed.
        """
        task = _make_task(
            title="系统测试",
            assignee="张三",
            deadline=(date.today() + timedelta(days=2)).strftime("%Y-%m-%d"),
            acceptance_criteria="测试计划、测试用例、测试执行、测试报告",
        )
        evidence = [
            "[source:测试计划] 系统测试计划已完成",
            "[source:周报] 测试用例编写中",
        ]
        ev = evaluate_with_rules(task, evidence)
        # Must NOT be completed
        assert ev.status != TaskStatus.COMPLETED
        assert ev.status != TaskStatus.MOSTLY_COMPLETED
        # Should be in_progress or needs_confirmation
        assert ev.status in (TaskStatus.IN_PROGRESS, TaskStatus.NEEDS_CONFIRMATION)


class TestGenerateWeeklyReport:
    """Test Markdown weekly report generation."""

    def test_generate_basic_report(self):
        evals = [
            _make_eval(TaskStatus.COMPLETED, "low", "任务A", "张三"),
            _make_eval(TaskStatus.IN_PROGRESS, "low", "任务B", "李四"),
            _make_eval(TaskStatus.DELAYED, "high", "任务C", "王五"),
        ]
        report = generate_weekly_report(evals, week_label="2026-W29", project_name="测试项目")
        assert "测试项目" in report
        assert "2026-W29" in report
        assert "任务A" in report
        assert "任务B" in report
        assert "任务C" in report
        assert "已完成" in report
        assert "已延期" in report
        assert "⚠️ 重点关注" in report

    def test_report_includes_high_risk_section(self):
        evals = [
            _make_eval(TaskStatus.DELAYED, "high", "高风险任务", "张三"),
        ]
        report = generate_weekly_report(evals)
        assert "⚠️ 重点关注" in report
        assert "高风险任务" in report

    def test_report_no_high_risk_section_when_none(self):
        evals = [
            _make_eval(TaskStatus.COMPLETED, "low", "正常任务", "张三"),
        ]
        report = generate_weekly_report(evals)
        assert "⚠️ 重点关注" not in report

    def test_report_without_project_name(self):
        evals = [_make_eval(TaskStatus.COMPLETED, "low")]
        report = generate_weekly_report(evals)
        assert "未命名项目" in report

    def test_report_status_counts(self):
        evals = [
            _make_eval(TaskStatus.COMPLETED, "low"),
            _make_eval(TaskStatus.COMPLETED, "low"),
            _make_eval(TaskStatus.IN_PROGRESS, "low"),
        ]
        report = generate_weekly_report(evals)
        assert "任务总数 | 3" in report
        assert "已完成 | 2" in report
        assert "进行中 | 1" in report

    def test_report_empty_evaluations(self):
        """Should handle empty evaluations list gracefully."""
        report = generate_weekly_report([])
        assert "任务总数 | 0" in report


class TestGenerateRiskCSV:
    """Test CSV risk list generation."""

    def test_generate_risk_csv_with_risks(self):
        evals = [
            _make_eval(TaskStatus.DELAYED, "high", "风险任务A", "张三"),
            _make_eval(TaskStatus.NEEDS_CONFIRMATION, "medium", "风险任务B", "李四"),
            _make_eval(TaskStatus.COMPLETED, "low", "正常任务", "王五"),
        ]
        csv_str = generate_risk_csv(evals)
        reader = csv.reader(io.StringIO(csv_str))
        rows = list(reader)
        # Header + 2 risk rows
        assert len(rows) >= 3
        assert rows[0] == [
            "task_title", "assignee", "status", "risk_level",
            "risk_reason", "recommendation", "deadline", "acceptance_criteria",
        ]
        titles = [r[0] for r in rows[1:]]
        # High risk first
        assert titles[0] == "风险任务A"
        assert titles[1] == "风险任务B"
        # Normal task excluded
        assert "正常任务" not in titles

    def test_generate_risk_csv_no_risks(self):
        evals = [_make_eval(TaskStatus.COMPLETED, "low")]
        csv_str = generate_risk_csv(evals)
        reader = csv.reader(io.StringIO(csv_str))
        rows = list(reader)
        # Only header, no risk rows
        assert len(rows) == 1


class TestGenerateNextWeekPlan:
    """Test next week plan generation."""

    def test_generate_plan_basic(self):
        evals = [
            _make_eval(TaskStatus.IN_PROGRESS, "low", "进行中任务", "张三"),
            _make_eval(TaskStatus.NOT_STARTED, "low", "待启动任务", "李四"),
        ]
        plan = generate_next_week_plan(evals, project_name="测试项目")
        assert "测试项目" in plan
        assert "草案" in plan
        assert "进行中任务" in plan
        assert "待启动任务" in plan
        assert "需要继续推进" in plan
        assert "需要启动" in plan

    def test_generate_plan_with_delayed(self):
        evals = [
            _make_eval(TaskStatus.DELAYED, "high", "逾期任务", "张三", deadline="2026-07-01"),
        ]
        plan = generate_next_week_plan(evals)
        assert "逾期任务" in plan
        assert "需立即处理" in plan

    def test_generate_plan_draft_only(self):
        """Plan must clearly state it is a draft and not overwrite original tasks."""
        evals = [_make_eval(TaskStatus.IN_PROGRESS, "low")]
        plan = generate_next_week_plan(evals)
        assert "草案" in plan
        assert "请勿直接覆盖原始任务列表" in plan

    def test_generate_plan_with_needs_confirmation(self):
        evals = [
            _make_eval(TaskStatus.NEEDS_CONFIRMATION, "medium", "待确认任务", "张三"),
        ]
        plan = generate_next_week_plan(evals)
        assert "待确认" in plan

    def test_generate_plan_with_mostly_completed(self):
        evals = [
            _make_eval(TaskStatus.MOSTLY_COMPLETED, "low", "基本完成任务", "张三"),
        ]
        plan = generate_next_week_plan(evals)
        assert "基本完成任务" in plan


class TestPromptAndParse:
    """Test LLM prompt building and response parsing."""

    def test_build_evaluation_prompt(self):
        from app.tools.task_checker import check_task

        task = _make_task(title="测试", assignee="A", deadline="2026-08-01")
        cr = check_task(task, ["已完成"])
        prompt = build_evaluation_prompt(cr)
        assert "测试" in prompt
        assert "A" in prompt

    def test_parse_valid_json(self):
        resp = '{"status": "completed", "risk_level": "low", "risk_reason": "一切正常"}'
        parsed = parse_llm_response(resp)
        assert parsed["status"] == "completed"

    def test_parse_json_with_markdown(self):
        resp = '```json\n{"status": "in_progress"}\n```'
        parsed = parse_llm_response(resp)
        assert parsed["status"] == "in_progress"

    def test_parse_invalid_json_returns_empty(self):
        resp = "not a json at all"
        parsed = parse_llm_response(resp)
        assert parsed == {}


# ===========================================================================
# Integration: full pipeline tests
# ===========================================================================


class TestFullPipeline:
    """End-to-end tests for the Phase C pipeline."""

    def test_full_pipeline_csv_to_reports(self):
        """Full pipeline: read CSV → evaluate → generate all reports."""
        csv_content = "任务名称,负责人,截止日期,优先级,验收标准\n"
        csv_content += "系统测试,张三,2026-08-01,high,测试计划和测试报告\n"
        csv_content += "接口开发,李四,2026-09-01,normal,API文档和测试通过\n"

        with tempfile.NamedTemporaryFile(
            mode="w", suffix=".csv", delete=False, encoding="utf-8-sig"
        ) as f:
            f.write(csv_content)
            tmp = f.name

        try:
            # Step 1: Read tasks
            tasks = read_tasks(tmp)
            assert len(tasks) == 2

            # Step 2: Evaluate with mock evidence
            evidence_map = {
                "系统测试": [
                    "[source:测试计划] 系统测试计划已完成",
                    "[source:周报] 测试用例编写中，部分已完成",
                ],
                "接口开发": [
                    "[source:周报] 接口开发已完成，API文档已交付",
                    "[source:测试报告] 接口测试全部通过",
                ],
            }
            evaluations = []
            for t in tasks:
                ev = evaluate_with_rules(t, evidence_map.get(t.title, []))
                evaluations.append(ev)

            # Step 3: Generate reports
            weekly = generate_weekly_report(evaluations, project_name="Demo项目")
            risk_csv = generate_risk_csv(evaluations)
            plan = generate_next_week_plan(evaluations, project_name="Demo项目")

            # Assertions
            assert weekly
            assert risk_csv
            assert plan
            assert "Demo项目" in weekly
            assert "Demo项目" in plan

            # 系统测试 should not be completed
            sys_test_ev = next(e for e in evaluations if e.task.title == "系统测试")
            assert sys_test_ev.status != TaskStatus.COMPLETED

        finally:
            os.unlink(tmp)

    def test_empty_evidence_all_not_started(self):
        """All tasks with empty evidence should be not_started."""
        csv_content = "任务名称,负责人\n任务A,张三\n任务B,李四\n"

        with tempfile.NamedTemporaryFile(
            mode="w", suffix=".csv", delete=False, encoding="utf-8-sig"
        ) as f:
            f.write(csv_content)
            tmp = f.name

        try:
            tasks = read_tasks(tmp)
            for t in tasks:
                ev = evaluate_with_rules(t, [])
                assert ev.status == TaskStatus.NOT_STARTED
        finally:
            os.unlink(tmp)


# ===========================================================================
# Edge case tests
# ===========================================================================


class TestEdgeCases:
    """Edge case and boundary tests."""

    def test_task_with_only_title_and_assignee(self):
        task = _make_task(title="最小任务", assignee="张三")
        cr = check_task(task, [])
        assert cr.rule_based_status == "not_started"

    def test_very_long_task_title(self):
        long_title = "这是一个非常非常长的任务标题" * 10
        task = _make_task(title=long_title, assignee="张三")
        cr = check_task(task, ["简短证据"])
        assert cr.task.title == long_title

    def test_special_chars_in_task_title(self):
        task = _make_task(title="任务<>&\"'测试", assignee="张三")
        cr = check_task(task, ["证据"])
        assert "任务<>&\"'测试" in cr.to_prompt_context()

    def test_task_with_multiline_acceptance_criteria(self):
        task = _make_task(
            title="复杂任务",
            assignee="张三",
            acceptance_criteria="需求评审通过；\n设计评审通过；\n代码合入主分支",
        )
        result = _match_accept_criteria(
            task.acceptance_criteria,
            "需求评审已通过，设计评审和代码合入跟进中",
        )
        # At least one should match via the new char-level matching
        matched_count = sum(1 for v in result.values() if v != "not_matched")
        assert matched_count >= 1


def _match_accept_criteria(criteria: str, evidence: str):
    """Helper for testing acceptance criteria matching.

    Delegates to the current implementation in task_checker so that the
    test helper stays in sync with the actual matching logic.
    """
    from app.tools.task_checker import _match_acceptance_criteria
    return _match_acceptance_criteria(criteria, evidence.lower())


# ===========================================================================
# Mock LLM client for testing LLM integration
# ===========================================================================


class _MockLLMClient:
    """Simulates the async LLM client interface for testing.

    Returns valid JSON or malformed text based on configuration.
    """

    def __init__(self, valid_json: dict | None = None, raise_on_call: bool = False):
        self._valid_json = valid_json
        self._raise_on_call = raise_on_call
        self.call_count = 0
        self.last_prompt = ""
        self.last_system_prompt = ""
        self.last_temperature = 0.0

    async def generate_text(
        self, prompt: str, system_prompt: str = "", temperature: float = 0.0
    ) -> str:
        self.call_count += 1
        self.last_prompt = prompt
        self.last_system_prompt = system_prompt
        self.last_temperature = temperature
        if self._raise_on_call:
            raise ConnectionError("LLM server unreachable (mock)")
        if self._valid_json is not None:
            import json as _json
            return _json.dumps(self._valid_json, ensure_ascii=False)
        return "invalid {{{ not json at all"


# ===========================================================================
# LLM explanation prompt tests
# ===========================================================================


class TestExplanationPrompt:
    """Tests for the Phase C explanation-only prompt builder."""

    def test_prompt_contains_rule_status(self):
        """Explanation prompt must state the rule-determined status as fixed."""
        task = _make_task(
            title="系统测试",
            assignee="张三",
            acceptance_criteria="测试计划和测试报告",
        )
        cr = check_task(task, ["[source:周报] 测试计划已完成，用例编写中"])
        prompt = build_explanation_prompt(cr)
        assert "系统测试" in prompt
        assert "张三" in prompt
        assert "测试计划已完成" in prompt
        # The prompt must declare the rule status
        assert cr.rule_based_status in prompt

    def test_prompt_does_not_ask_llm_to_change_status(self):
        """The LLM is NOT asked to decide status — rules already did."""
        task = _make_task(title="接口开发", assignee="李四",
                          acceptance_criteria="API文档和测试通过")
        cr = check_task(task, ["API文档已交付，测试全部通过"])
        prompt = build_explanation_prompt(cr)
        # Prompt should not present status options for the LLM to choose
        assert "completed（已完成）" not in prompt
        assert "可选的评估状态" not in prompt

    def test_prompt_empty_evidence(self):
        """Prompt works with no evidence (not_started tasks)."""
        task = _make_task(title="未开始任务", assignee="张三")
        cr = check_task(task, [])
        prompt = build_explanation_prompt(cr)
        assert "未开始任务" in prompt
        assert "（无证据）" in prompt

    def test_prompt_missing_ac_detail(self):
        """Prompt works when acceptance criteria matching is empty."""
        task = _make_task(title="简单任务", assignee="张三")
        cr = check_task(task, ["已完成所有工作"])
        prompt = build_explanation_prompt(cr)
        assert "（无验收标准）" in prompt


# ===========================================================================
# LLM integration tests (async)
# ===========================================================================


class TestEvaluateWithLLM:
    """Tests for evaluate_with_llm using a mock LLM client."""

    @pytest.mark.asyncio
    async def test_llm_explanation_applied(self):
        """LLM explanation overrides evidence_summary and risk_reason."""
        task = _make_task(
            title="接口开发",
            assignee="李四",
            acceptance_criteria="API文档和测试通过",
            deadline="2026-08-01",
        )
        evidence = ["API文档已交付", "测试全部通过"]

        mock_llm = _MockLLMClient(valid_json={
            "explanation": "验收标准全部满足，交付物齐全，判断为已完成",
            "risk_reason": "无风险，任务按计划推进并已验收交付",
            "recommendation": "更新时间表，标记任务完成",
        })

        ev = await evaluate_with_llm(task, evidence, mock_llm)
        assert ev.status == TaskStatus.COMPLETED  # status from rules, NOT llm
        assert ev.evidence_summary == "验收标准全部满足，交付物齐全，判断为已完成"
        assert ev.risk_reason == "无风险，任务按计划推进并已验收交付"
        assert ev.recommendation == "更新时间表，标记任务完成"
        assert mock_llm.call_count == 1

    @pytest.mark.asyncio
    async def test_llm_status_not_overridden(self):
        """LLM does NOT override the rule-determined status."""
        task = _make_task(
            title="系统测试",
            assignee="周八",
            acceptance_criteria="测试计划和测试报告",
            deadline="2026-07-22",
        )
        evidence = ["测试计划已完成，部分用例编写中"]

        # Mock LLM returning valid explanation
        mock_llm = _MockLLMClient(valid_json={
            "explanation": "看起来基本完成了",
            "risk_reason": "有点风险",
            "recommendation": "再看看",
        })

        ev = await evaluate_with_llm(task, evidence, mock_llm)
        # Status must still be rule-based (needs_confirmation/in_progress),
        # not changed by LLM to "completed"
        assert ev.status != TaskStatus.COMPLETED

    @pytest.mark.asyncio
    async def test_llm_prompt_receives_rule_result(self):
        """The LLM receives the rule-determined status in the prompt."""
        task = _make_task(
            title="UI开发",
            assignee="李四",
            acceptance_criteria="代码评审通过、文档完成",
            deadline="2026-08-01",
        )
        evidence = ["代码已完成，文档还在整理中"]
        mock_llm = _MockLLMClient(valid_json={
            "explanation": "大部分完成",
            "risk_reason": "文档不完整",
            "recommendation": "补全文档",
        })

        await evaluate_with_llm(task, evidence, mock_llm)
        prompt = mock_llm.last_prompt
        # Rule-determined status must appear in the prompt
        assert "状态:" in prompt
        # Task info must be in the prompt
        assert "UI开发" in prompt

    @pytest.mark.asyncio
    async def test_llm_fallback_on_invalid_json(self):
        """When LLM returns unparseable JSON, fall back to rule evaluation."""
        task = _make_task(title="接口开发", assignee="李四",
                          acceptance_criteria="API文档和测试通过")
        evidence = ["API文档已交付，测试全部通过"]

        mock_llm = _MockLLMClient(valid_json=None)  # will return invalid text
        ev = await evaluate_with_llm(task, evidence, mock_llm)

        # Status should still be correct (rules worked), but explanation
        # should be the rule-based one
        assert ev.status == TaskStatus.COMPLETED
        # Rule evidence summary is still present (fallback succeeded)
        assert ev.evidence_summary

    @pytest.mark.asyncio
    async def test_llm_fallback_on_connection_error(self):
        """When LLM call raises an exception, fall back gracefully."""
        task = _make_task(title="接口开发", assignee="李四",
                          acceptance_criteria="API文档和测试通过")
        evidence = ["API文档已交付，测试全部通过"]

        mock_llm = _MockLLMClient(raise_on_call=True)
        ev = await evaluate_with_llm(task, evidence, mock_llm)

        # Should not crash, should return rule evaluation
        assert ev.status == TaskStatus.COMPLETED
        assert ev.evidence_summary  # rule-based fallback

    @pytest.mark.asyncio
    async def test_llm_partial_json_applied(self):
        """LLM response with partial fields only overrides what is present."""
        task = _make_task(title="接口开发", assignee="李四",
                          acceptance_criteria="API文档和测试通过")
        evidence = ["API文档已交付，测试全部通过"]

        mock_llm = _MockLLMClient(valid_json={
            "explanation": "LLM解释",
            # No risk_reason or recommendation in response
        })

        ev = await evaluate_with_llm(task, evidence, mock_llm)
        assert ev.evidence_summary == "LLM解释"
        # risk_reason should still have the rule-based value
        assert ev.risk_reason

    @pytest.mark.asyncio
    async def test_llm_not_called_for_not_started(self):
        """LLM is still called for not_started tasks (to explain why)."""
        task = _make_task(title="性能优化", assignee="张三")
        evidence: list[str] = []

        mock_llm = _MockLLMClient(valid_json={
            "explanation": "无任何交付物证据",
            "risk_reason": "尚未开始",
            "recommendation": "尽快启动",
        })

        ev = await evaluate_with_llm(task, evidence, mock_llm)
        assert ev.status == TaskStatus.NOT_STARTED
        assert mock_llm.call_count == 1


class TestBatchEvaluateWithLLM:
    """Tests for batch LLM evaluation."""

    @pytest.mark.asyncio
    async def test_batch_sequential(self):
        """All tasks processed, one LLM call per task."""
        tasks = [
            _make_task(title="任务A", assignee="张三",
                       acceptance_criteria="需求评审通过"),
            _make_task(title="任务B", assignee="李四",
                       acceptance_criteria="API文档和测试通过"),
        ]
        evidence_map = {
            "任务A": ["需求评审已完成"],
            "任务B": [],
        }
        mock_llm = _MockLLMClient(valid_json={
            "explanation": "ok", "risk_reason": "ok", "recommendation": "ok",
        })

        evals = await batch_evaluate_with_llm(tasks, evidence_map, mock_llm)
        assert len(evals) == 2
        assert mock_llm.call_count == 2

    @pytest.mark.asyncio
    async def test_batch_mixed_fallback(self):
        """One task has evidence, one doesn't — both evaluated correctly."""
        tasks = [
            _make_task(title="有证据任务", assignee="张三"),
            _make_task(title="无证据任务", assignee="李四"),
        ]
        evidence_map = {
            "有证据任务": ["已完成"],
            "无证据任务": [],
        }
        mock_llm = _MockLLMClient(valid_json={
            "explanation": "ok", "risk_reason": "ok", "recommendation": "ok",
        })

        evals = await batch_evaluate_with_llm(tasks, evidence_map, mock_llm)
        assert evals[0].status != TaskStatus.NOT_STARTED  # has evidence
        assert evals[1].status == TaskStatus.NOT_STARTED    # no evidence

    @pytest.mark.asyncio
    async def test_batch_empty(self):
        """Empty task list returns empty evaluations."""
        mock_llm = _MockLLMClient()
        evals = await batch_evaluate_with_llm([], {}, mock_llm)
        assert len(evals) == 0
        assert mock_llm.call_count == 0


# ===========================================================================
# Full pipeline with LLM (verify_phase_c style)
# ===========================================================================


class TestFullPipelineWithLLM:
    """End-to-end tests with LLM integration."""

    @pytest.mark.asyncio
    async def test_system_test_still_not_completed_with_llm(self):
        """Critical acceptance: 系统测试 with only plan + partial records
        must not be marked completed by the LLM."""
        task = _make_task(
            title="系统测试",
            assignee="周八",
            deadline="2026-07-22",
            acceptance_criteria="测试计划、测试报告、覆盖率达标",
        )
        evidence = [
            "[source:测试计划] 系统测试计划已完成",
            "[source:周报] 测试用例编写中，部分模块已通过",
        ]

        mock_llm = _MockLLMClient(valid_json={
            "explanation": "测试计划已完成，但有部分用例还在编写中，没有最终报告",
            "risk_reason": "缺少最终测试报告，覆盖率不明确",
            "recommendation": "尽快完成剩余用例并输出测试报告",
        })

        ev = await evaluate_with_llm(task, evidence, mock_llm)

        # MUST NOT be completed
        assert ev.status != TaskStatus.COMPLETED
        # Status should reflect the incomplete evidence
        # The fixed historical deadline is intentionally overdue.  Delayed is
        # therefore the expected safe classification as well as the two
        # incomplete-but-not-overdue classifications.
        assert ev.status in (
            TaskStatus.NEEDS_CONFIRMATION,
            TaskStatus.IN_PROGRESS,
            TaskStatus.DELAYED,
        )

    @pytest.mark.asyncio
    async def test_completed_task_with_llm_explanation(self):
        """Fully completed task gets LLM explanation."""
        task = _make_task(
            title="接口开发",
            assignee="李四",
            deadline="2026-08-10",
            acceptance_criteria="API文档完成；测试全部通过；代码合入主分支",
        )
        evidence = [
            "[source:周报] API文档已完成并交付",
            "[source:测试报告] 接口测试全部通过，覆盖率95%",
            "[source:Git] 代码PR #456 已合入主分支",
        ]

        mock_llm = _MockLLMClient(valid_json={
            "explanation": "三条验收标准全部满足：API文档已交付，测试全部通过，代码已合入主分支",
            "risk_reason": "无风险，任务圆满完成",
            "recommendation": "标记任务为已完成，关闭相关工单",
        })

        ev = await evaluate_with_llm(task, evidence, mock_llm)
        assert ev.status == TaskStatus.COMPLETED
        assert "LLM" not in ev.status.value  # status is rule-based

    @pytest.mark.asyncio
    async def test_cancelled_with_llm(self):
        """Cancelled task gets LLM risk explanation."""
        task = _make_task(
            title="废弃功能",
            assignee="吴九",
            deadline="2026-07-31",
            acceptance_criteria="功能已废弃下线",
        )
        evidence = ["该功能已废弃，不再维护，计划下版本移除"]

        mock_llm = _MockLLMClient(valid_json={
            "explanation": "任务明确标注为废弃，已不再维护",
            "risk_reason": "低风险，但需确认是否有下游依赖",
            "recommendation": "确认无下游依赖后安全下线",
        })

        ev = await evaluate_with_llm(task, evidence, mock_llm)
        assert ev.status == TaskStatus.CANCELLED

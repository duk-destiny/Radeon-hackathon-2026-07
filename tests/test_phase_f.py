"""Phase F — Task lifecycle & human confirmation tests.

Coverage
--------
* Task CRUD (create, read, update, list) via API and service
* Task state machine (all valid transitions, invalid transitions, history)
* Candidate extraction from meeting notes
* Human confirmation queue (accept / modify / ignore)
* CSV import with diff preview, dedup, and confirmation
* XLSX import with diff preview, dedup, and confirmation
* Operation audit trail
* Report prioritization of task DB over CSV/XLSX
* Edge cases: not found, already processed, invalid transitions
"""

from __future__ import annotations

import io
import os
from datetime import date, datetime
from pathlib import Path

from fastapi.testclient import TestClient

from app.config import Settings
from app.main import create_app
from app.schemas.models import (
    ConfirmationAction,
    PhaseFTaskStatus,
    TaskCreate,
    TaskImportConfirm,
    TaskStatusTransition,
    TaskUpdate,
)
from app.schemas.task_sql import TASK_STATUSES, is_valid_transition
from app.services.task_lifecycle import TaskLifecycleService, _parse_csv, _parse_xlsx


# ---------------------------------------------------------------------------
# helpers
# ---------------------------------------------------------------------------


def _settings(tmp_path: Path) -> Settings:
    return Settings(
        project_root=tmp_path / "projects",
        output_root=tmp_path / "outputs",
        vector_db_root=tmp_path / "vectors",
        sqlite_path=tmp_path / "sqlite" / "projectpack.db",
        log_root=tmp_path / "logs",
    )


def _setup_project(client: TestClient) -> None:
    assert client.post(
        "/api/projects", json={"project_id": "demo-project", "name": "Demo"}
    ).status_code == 201


def _create_task(client: TestClient, title: str = "Test task", status: str = "pending_confirmation") -> dict:
    return client.post(
        "/api/projects/demo-project/tasks",
        json={
            "title": title,
            "owner": "alice",
            "due_date": "2026-08-15",
            "priority": "high",
            "acceptance_criteria": "All tests pass",
            "dependencies": ["task-abc"],
            "source_ref": "weekly_plan.md",
            "status": status,
        },
    ).json()


# ============================================================================
# Task status state machine (unit)
# ============================================================================


def test_state_machine_allowed_transitions() -> None:
    """All documented transitions are valid."""
    assert is_valid_transition("pending_confirmation", "not_started") is True
    assert is_valid_transition("pending_confirmation", "cancelled") is True
    assert is_valid_transition("not_started", "in_progress") is True
    assert is_valid_transition("not_started", "cancelled") is True
    assert is_valid_transition("in_progress", "completed") is True
    assert is_valid_transition("in_progress", "mostly_completed") is True
    assert is_valid_transition("in_progress", "delayed") is True
    assert is_valid_transition("mostly_completed", "completed") is True
    assert is_valid_transition("delayed", "in_progress") is True
    assert is_valid_transition("delayed", "completed") is True
    assert is_valid_transition("delayed", "cancelled") is True


def test_state_machine_invalid_transitions() -> None:
    """Cannot go from cancelled back to completed or any other state."""
    assert is_valid_transition("cancelled", "completed") is False
    assert is_valid_transition("cancelled", "in_progress") is False
    assert is_valid_transition("cancelled", "not_started") is False
    assert is_valid_transition("completed", "cancelled") is False
    assert is_valid_transition("completed", "in_progress") is False


def test_all_statuses_in_allowed_transitions() -> None:
    """Every known status appears somewhere in the transition map."""
    for status in TASK_STATUSES:
        assert status in TASK_STATUSES  # it's a member


# ============================================================================
# Task CRUD — API
# ============================================================================


def test_create_and_read_task(tmp_path: Path) -> None:
    """Create a task and read it back via the API."""
    app = create_app(_settings(tmp_path))
    with TestClient(app) as client:
        _setup_project(client)

        created = _create_task(client, "Review API docs")
        assert created["title"] == "Review API docs"
        assert created["owner"] == "alice"
        assert created["priority"] == "high"
        assert "id" in created

        # read back
        fetched = client.get(f"/api/projects/demo-project/tasks/{created['id']}").json()
        assert fetched["title"] == "Review API docs"
        assert fetched["id"] == created["id"]


def test_list_tasks(tmp_path: Path) -> None:
    """List all tasks for a project, filtered by status."""
    app = create_app(_settings(tmp_path))
    with TestClient(app) as client:
        _setup_project(client)

        t1 = _create_task(client, "Task A")
        t2 = _create_task(client, "Task B")

        all_tasks = client.get("/api/projects/demo-project/tasks").json()
        assert len(all_tasks) >= 2

        # filter by status
        pending = client.get("/api/projects/demo-project/tasks?status=pending_confirmation").json()
        assert any(t["id"] == t1["id"] for t in pending)


def test_update_task(tmp_path: Path) -> None:
    """Patch a task with updated fields."""
    app = create_app(_settings(tmp_path))
    with TestClient(app) as client:
        _setup_project(client)
        created = _create_task(client, "Old title")

        updated = client.patch(
            f"/api/projects/demo-project/tasks/{created['id']}",
            json={"title": "New title", "priority": "critical"},
        ).json()
        assert updated["title"] == "New title"
        assert updated["priority"] == "critical"


def test_get_task_404(tmp_path: Path) -> None:
    """Non-existent task returns 404."""
    app = create_app(_settings(tmp_path))
    with TestClient(app) as client:
        _setup_project(client)
        resp = client.get("/api/projects/demo-project/tasks/nonexistent-id")
        assert resp.status_code == 404


def test_update_task_404(tmp_path: Path) -> None:
    """Update non-existent task returns 404."""
    app = create_app(_settings(tmp_path))
    with TestClient(app) as client:
        _setup_project(client)
        resp = client.patch(
            "/api/projects/demo-project/tasks/nonexistent-id",
            json={"title": "Nope"},
        )
        assert resp.status_code == 404


# ============================================================================
# State machine — API
# ============================================================================


def test_transition_task_valid(tmp_path: Path) -> None:
    """Transition a task from pending_confirmation to not_started."""
    app = create_app(_settings(tmp_path))
    with TestClient(app) as client:
        _setup_project(client)

        # create with pending_confirmation then accept via transition
        created = _create_task(client, "Task", status="not_started")
        # not_started -> in_progress
        result = client.post(
            f"/api/projects/demo-project/tasks/{created['id']}/transition",
            json={"status": "in_progress", "reason": "Started work", "changed_by": "alice"},
        ).json()
        assert result["status"] == "in_progress"


def test_transition_invalid_returns_400(tmp_path: Path) -> None:
    """Cannot jump from pending_confirmation directly to completed."""
    app = create_app(_settings(tmp_path))
    with TestClient(app) as client:
        _setup_project(client)
        created = _create_task(client, "Task")

        resp = client.post(
            f"/api/projects/demo-project/tasks/{created['id']}/transition",
            json={"status": "completed", "reason": "Nope", "changed_by": "alice"},
        )
        assert resp.status_code == 400


def test_task_history(tmp_path: Path) -> None:
    """Task history returns all status change events."""
    app = create_app(_settings(tmp_path))
    with TestClient(app) as client:
        _setup_project(client)

        created = _create_task(client, "History task", status="not_started")
        client.post(
            f"/api/projects/demo-project/tasks/{created['id']}/transition",
            json={"status": "in_progress", "reason": "start", "changed_by": "alice"},
        )
        client.post(
            f"/api/projects/demo-project/tasks/{created['id']}/transition",
            json={"status": "completed", "reason": "done", "changed_by": "alice"},
        )

        history = client.get(f"/api/projects/demo-project/tasks/{created['id']}/history").json()
        assert len(history) >= 3


def test_cannot_transition_from_cancelled(tmp_path: Path) -> None:
    """Cancelled task cannot be reactivated."""
    app = create_app(_settings(tmp_path))
    with TestClient(app) as client:
        _setup_project(client)

        created = _create_task(client, "Doomed", status="not_started")
        # cancel
        client.post(
            f"/api/projects/demo-project/tasks/{created['id']}/transition",
            json={"status": "cancelled", "reason": "No longer needed", "changed_by": "alice"},
        )
        # try to reactivate
        resp = client.post(
            f"/api/projects/demo-project/tasks/{created['id']}/transition",
            json={"status": "in_progress", "reason": "Actually needed", "changed_by": "alice"},
        )
        assert resp.status_code == 400


# ============================================================================
# Candidate extraction
# ============================================================================


def test_extract_candidates_from_meeting_notes(tmp_path: Path) -> None:
    """Extract candidate tasks from meeting-note-style text."""
    app = create_app(_settings(tmp_path))
    with TestClient(app) as client:
        _setup_project(client)

        source = (
            "1. @alice needs to review the API docs by 2026-08-01\n"
            "2. @bob should update the deployment pipeline\n"
            "3. Schedule a team retro\n"
            "TODO: Fix the login bug by 2026-07-30\n"
        )
        result = client.post(
            "/api/projects/demo-project/tasks/extract",
            json={"source_text": source, "source_kind": "meeting_notes", "project_id": "demo-project"},
        ).json()
        assert "candidates" in result
        assert len(result["candidates"]) > 0
        # highest confidence candidate should have owner and deadline
        assert any(c.get("owner") == "alice" for c in result["candidates"])
        assert any(
            c.get("title") and "review" in c.get("title", "").lower()
            for c in result["candidates"]
        )


def test_submit_candidates_to_queue(tmp_path: Path) -> None:
    """Submit extracted candidates to the confirmation queue."""
    app = create_app(_settings(tmp_path))
    with TestClient(app) as client:
        _setup_project(client)

        candidates = [
            {
                "title": "Review docs",
                "owner": "alice",
                "source_kind": "meeting_notes",
                "confidence": 0.8,
            }
        ]
        result = client.post(
            "/api/projects/demo-project/tasks/submit-candidates",
            json={"candidates": candidates},
        ).json()
        assert len(result) == 1
        assert result[0]["status"] == "pending"
        assert result[0]["candidate_title"] == "Review docs"


# ============================================================================
# Human confirmation queue
# ============================================================================


def test_confirmation_queue_list(tmp_path: Path) -> None:
    """List items in the confirmation queue."""
    app = create_app(_settings(tmp_path))
    with TestClient(app) as client:
        _setup_project(client)

        # create task with pending_confirmation (goes into queue)
        created = _create_task(client, "Needs approval", status="pending_confirmation")

        queue = client.get("/api/projects/demo-project/tasks/confirmation-queue").json()
        assert len(queue) >= 1
        assert any(c["task_id"] == created["id"] for c in queue)


def test_confirm_accept(tmp_path: Path) -> None:
    """Accept a pending confirmation — task moves to not_started."""
    app = create_app(_settings(tmp_path))
    with TestClient(app) as client:
        _setup_project(client)
        created = _create_task(client, "Accept me", status="pending_confirmation")

        result = client.post(
            f"/api/projects/demo-project/tasks/confirmation/{created['id']}",
            json={
                "action": "accept",
                "confirmed_by": "alice",
                "confirmation_basis": "Approved in weekly",
                "confirmation_notes": "Priority item",
            },
        ).json()

        assert result["status"] == "accepted"
        assert result["confirmed_by"] == "alice"
        assert result["confirmation_basis"] == "Approved in weekly"

        # task should now be not_started
        task = client.get(f"/api/projects/demo-project/tasks/{created['id']}").json()
        assert task["status"] == "not_started"


def test_confirm_ignore(tmp_path: Path) -> None:
    """Ignore a pending confirmation — task moves to cancelled."""
    app = create_app(_settings(tmp_path))
    with TestClient(app) as client:
        _setup_project(client)
        created = _create_task(client, "Ignore me", status="pending_confirmation")

        result = client.post(
            f"/api/projects/demo-project/tasks/confirmation/{created['id']}",
            json={
                "action": "ignore",
                "confirmed_by": "alice",
                "confirmation_basis": "Out of scope",
                "confirmation_notes": "Will reconsider next quarter",
            },
        ).json()

        assert result["status"] == "ignored"

        # task should now be cancelled
        task = client.get(f"/api/projects/demo-project/tasks/{created['id']}").json()
        assert task["status"] == "cancelled"


def test_confirm_modify(tmp_path: Path) -> None:
    """Modify a pending confirmation with new fields then accept."""
    app = create_app(_settings(tmp_path))
    with TestClient(app) as client:
        _setup_project(client)
        created = _create_task(client, "Modify me", status="pending_confirmation")

        result = client.post(
            f"/api/projects/demo-project/tasks/confirmation/{created['id']}",
            json={
                "action": "modify",
                "confirmed_by": "alice",
                "confirmation_basis": "Scope changed",
                "confirmation_notes": "Narrowed scope",
                "modified_title": "Modified task title",
                "modified_priority": "low",
            },
        ).json()

        assert result["status"] == "accepted"

        # task should have the modified fields
        task = client.get(f"/api/projects/demo-project/tasks/{created['id']}").json()
        assert task["title"] == "Modified task title"
        assert task["priority"] == "low"
        assert task["status"] == "not_started"


def test_confirm_already_processed_returns_400(tmp_path: Path) -> None:
    """Cannot process a confirmation that was already handled."""
    app = create_app(_settings(tmp_path))
    with TestClient(app) as client:
        _setup_project(client)
        created = _create_task(client, "Onetime", status="pending_confirmation")

        # first accept
        client.post(
            f"/api/projects/demo-project/tasks/confirmation/{created['id']}",
            json={"action": "accept", "confirmed_by": "alice"},
        )
        # second attempt
        resp = client.post(
            f"/api/projects/demo-project/tasks/confirmation/{created['id']}",
            json={"action": "accept", "confirmed_by": "alice"},
        )
        assert resp.status_code == 400


def test_confirm_not_found_returns_404(tmp_path: Path) -> None:
    """Processing a non-existent confirmation returns 404."""
    app = create_app(_settings(tmp_path))
    with TestClient(app) as client:
        _setup_project(client)
        resp = client.post(
            "/api/projects/demo-project/tasks/confirmation/nonexistent-id",
            json={"action": "accept", "confirmed_by": "alice"},
        )
        assert resp.status_code == 404


# ============================================================================
# CSV import with diff preview + dedup + confirmation
# ============================================================================


def test_csv_import_preview(tmp_path: Path) -> None:
    """Preview a CSV import returning diff counts."""
    app = create_app(_settings(tmp_path))
    with TestClient(app) as client:
        _setup_project(client)

        csv_content = (
            "title,owner,due_date,priority,acceptance_criteria\n"
            "Task Alpha,alice,2026-08-01,high,Must pass QA\n"
            "Task Beta,bob,2026-08-15,medium,Approved by PM\n"
        )
        files = {"file": ("tasks.csv", io.BytesIO(csv_content.encode("utf-8")), "text/csv")}
        resp = client.post(
            "/api/projects/demo-project/tasks/import-preview", files=files
        )
        assert resp.status_code == 200
        data = resp.json()
        assert data["new_rows"] == 2
        assert data["duplicate_rows"] == 0
        assert data["conflict_rows"] == 0


def test_csv_import_confirm(tmp_path: Path) -> None:
    """Confirm and execute a CSV import."""
    app = create_app(_settings(tmp_path))
    with TestClient(app) as client:
        _setup_project(client)

        csv_content = (
            "title,owner,due_date,priority,acceptance_criteria\n"
            "Import Task,alice,2026-08-20,high,Must work\n"
        )
        files = {"file": ("import.csv", io.BytesIO(csv_content.encode("utf-8")), "text/csv")}
        data = {"confirmed_by": "alice", "skip_duplicates": "true", "overwrite_conflicts": "false"}

        resp = client.post(
            "/api/projects/demo-project/tasks/import-confirm",
            files=files,
            data=data,
        )
        assert resp.status_code == 201
        result = resp.json()
        assert result["imported"] >= 1
        assert result["errors"] == 0

        # imported tasks should appear in the task list
        all_tasks = client.get("/api/projects/demo-project/tasks").json()
        assert any(t["title"] == "Import Task" for t in all_tasks)


def test_csv_import_dedup(tmp_path: Path) -> None:
    """Re-importing the same CSV deduplicates by fingerprint."""
    app = create_app(_settings(tmp_path))
    with TestClient(app) as client:
        _setup_project(client)

        csv_content = (
            "title,owner,due_date,priority\n"
            "Unique Task,alice,2026-09-01,high\n"
        )

        # first import
        files1 = {"file": ("tasks.csv", io.BytesIO(csv_content.encode("utf-8")), "text/csv")}
        data1 = {"confirmed_by": "alice", "skip_duplicates": "true", "overwrite_conflicts": "false"}
        r1 = client.post("/api/projects/demo-project/tasks/import-confirm", files=files1, data=data1)
        assert r1.status_code == 201
        assert r1.json()["imported"] == 1

        # second import — should be deduped
        files2 = {"file": ("tasks2.csv", io.BytesIO(csv_content.encode("utf-8")), "text/csv")}
        data2 = {"confirmed_by": "alice", "skip_duplicates": "true", "overwrite_conflicts": "false"}
        r2 = client.post("/api/projects/demo-project/tasks/import-confirm", files=files2, data=data2)
        assert r2.status_code == 201
        assert r2.json()["imported"] == 0  # deduped
        assert r2.json()["skipped"] == 1


def test_csv_import_preview_duplicates(tmp_path: Path) -> None:
    """Preview shows duplicates after a previous import."""
    app = create_app(_settings(tmp_path))
    with TestClient(app) as client:
        _setup_project(client)

        csv_content = (
            "title,owner,due_date,priority\n"
            "Dup Task,alice,2026-10-01,high\n"
        )

        # do the import first
        files = {"file": ("tasks.csv", io.BytesIO(csv_content.encode("utf-8")), "text/csv")}
        client.post(
            "/api/projects/demo-project/tasks/import-confirm",
            files=files,
            data={"confirmed_by": "alice", "skip_duplicates": "true", "overwrite_conflicts": "false"},
        )

        # preview again
        files2 = {"file": ("tasks2.csv", io.BytesIO(csv_content.encode("utf-8")), "text/csv")}
        preview = client.post("/api/projects/demo-project/tasks/import-preview", files=files2).json()
        assert preview["duplicate_rows"] == 1
        assert preview["new_rows"] == 0


def test_import_preview_unsupported_format(tmp_path: Path) -> None:
    """Unsupported file format returns 400."""
    app = create_app(_settings(tmp_path))
    with TestClient(app) as client:
        _setup_project(client)

        files = {"file": ("doc.txt", io.BytesIO(b"not a csv"), "text/plain")}
        resp = client.post("/api/projects/demo-project/tasks/import-preview", files=files)
        assert resp.status_code == 400


# ============================================================================
# XLSX import
# ============================================================================


def test_xlsx_import_preview(tmp_path: Path) -> None:
    """Preview an XLSX import."""
    try:
        from openpyxl import Workbook
    except ImportError:
        import pytest
        pytest.skip("openpyxl not installed")

    app = create_app(_settings(tmp_path))
    with TestClient(app) as client:
        _setup_project(client)

        wb = Workbook()
        ws = wb.active
        ws.append(["title", "owner", "due_date", "priority", "acceptance_criteria"])
        ws.append(["XLSX Task", "alice", "2026-08-01", "high", "Must pass"])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        files = {"file": ("tasks.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        resp = client.post("/api/projects/demo-project/tasks/import-preview", files=files)
        assert resp.status_code == 200
        data = resp.json()
        assert data["new_rows"] >= 1


def test_xlsx_import_confirm(tmp_path: Path) -> None:
    """Confirm and execute an XLSX import."""
    try:
        from openpyxl import Workbook
    except ImportError:
        import pytest
        pytest.skip("openpyxl not installed")

    app = create_app(_settings(tmp_path))
    with TestClient(app) as client:
        _setup_project(client)

        wb = Workbook()
        ws = wb.active
        ws.append(["title", "owner", "due_date", "priority"])
        ws.append(["XLSX Confirm", "bob", "2026-09-15", "medium"])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        files = {"file": ("tasks.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        data = {"confirmed_by": "bob", "skip_duplicates": "true", "overwrite_conflicts": "false"}
        resp = client.post("/api/projects/demo-project/tasks/import-confirm", files=files, data=data)
        assert resp.status_code == 201
        result = resp.json()
        assert result["imported"] >= 1
        assert result["errors"] == 0


# ============================================================================
# Operation audit trail
# ============================================================================


def test_operation_audit_log(tmp_path: Path) -> None:
    """Audit log records task operations."""
    app = create_app(_settings(tmp_path))
    with TestClient(app) as client:
        _setup_project(client)

        # create a few tasks
        _create_task(client, "Audit task 1")
        _create_task(client, "Audit task 2")

        audit = client.get("/api/projects/demo-project/tasks/audit-log").json()
        assert len(audit) >= 2
        assert all("entity_type" in a for a in audit)
        assert all("operation" in a for a in audit)
        assert any(a["operation"] == "create" for a in audit)


# ============================================================================
# Report integration: task DB priority over CSV/XLSX
# ============================================================================


def test_load_tasks_prioritizes_db_over_csv(tmp_path: Path) -> None:
    """When task DB has confirmed tasks, load_tasks uses those, not CSV."""
    from app.services.phase_c import load_tasks

    settings = _settings(tmp_path)
    project_id = "demo-db-first"
    source = settings.project_root / project_id / "source"
    source.mkdir(parents=True)
    (source / "tasks.csv").write_text(
        "title,assignee,deadline,priority,acceptance_criteria\n"
        "CSV Task,bob,2026-12-01,low,from csv\n",
        encoding="utf-8",
    )

    # seed the task DB with confirmed tasks
    from app.services.task_lifecycle import TaskLifecycleService
    db_path = Path(settings.sqlite_path) / "projects" / project_id / "tasks.db"
    svc = TaskLifecycleService(db_path)
    svc.create_task(
        project_id,
        TaskCreate(
            title="DB Task Alpha",
            owner="alice",
            due_date=date(2026, 8, 1),
            priority="high",
            acceptance_criteria="from db",
            status="not_started",
        ),
    )
    svc.create_task(
        project_id,
        TaskCreate(
            title="DB Task Beta",
            owner="bob",
            due_date=date(2026, 9, 1),
            priority="medium",
            status="in_progress",
        ),
    )

    tasks = load_tasks(project_id, settings=settings)
    # should prioritize DB tasks
    assert len(tasks) >= 2
    assert any(t.title == "DB Task Alpha" for t in tasks)
    assert any(t.title == "DB Task Beta" for t in tasks)
    # the CSV task should NOT appear when DB tasks exist
    assert not any(t.title == "CSV Task" for t in tasks)


def test_load_tasks_falls_back_to_csv_when_db_empty(tmp_path: Path) -> None:
    """When task DB exists but has no confirmed tasks, fall back to CSV."""
    from app.services.phase_c import load_tasks

    settings = _settings(tmp_path)
    project_id = "demo-fallback"
    source = settings.project_root / project_id / "source"
    source.mkdir(parents=True)
    (source / "tasks.csv").write_text(
        "title,assignee,deadline,priority,acceptance_criteria\n"
        "Fallback Task,bob,2026-12-15,low,from csv\n",
        encoding="utf-8",
    )

    # create an empty DB
    from app.services.task_lifecycle import TaskLifecycleService
    db_path = Path(settings.sqlite_path) / "projects" / project_id / "tasks.db"
    TaskLifecycleService(db_path)

    tasks = load_tasks(project_id, settings=settings)
    # DB has no confirmed tasks → falls back to CSV
    assert len(tasks) == 1
    assert tasks[0].title == "Fallback Task"


# ============================================================================
# Service-layer unit tests
# ============================================================================


def test_service_create_and_list(tmp_path: Path) -> None:
    """Direct service call: create and list tasks."""
    db = tmp_path / "test_project.db"
    svc = TaskLifecycleService(db)

    task = svc.create_task(
        "proj-1",
        TaskCreate(
            title="Service task",
            owner="charlie",
            due_date=date(2026, 7, 30),
            priority="high",
            acceptance_criteria="Unit tested",
            status="not_started",
        ),
    )
    assert task.title == "Service task"
    assert task.status == "not_started"

    tasks = svc.list_tasks("proj-1")
    assert len(tasks) == 1
    assert tasks[0].id == task.id


def test_service_update_task(tmp_path: Path) -> None:
    """Direct service call: update task fields."""
    db = tmp_path / "test_update.db"
    svc = TaskLifecycleService(db)

    task = svc.create_task(
        "proj-update",
        TaskCreate(title="Old name", status="not_started"),
    )
    updated = svc.update_task(
        "proj-update",
        task.id,
        TaskUpdate(title="New name", priority="critical"),
    )
    assert updated.title == "New name"
    assert updated.priority == "critical"


def test_service_state_machine(tmp_path: Path) -> None:
    """Direct service call: state transitions."""
    db = tmp_path / "test_state.db"
    svc = TaskLifecycleService(db)

    task = svc.create_task(
        "proj-state",
        TaskCreate(title="State test", status="not_started"),
    )

    svc.transition_status(
        "proj-state",
        task.id,
        TaskStatusTransition(status="in_progress", reason="Working on it"),
    )

    updated = svc.get_task("proj-state", task.id)
    assert updated.status == "in_progress"

    history = svc.get_task_history("proj-state", task.id)
    assert len(history) >= 2
    assert history[-1].to_status == "in_progress"


def test_service_invalid_transition_raises(tmp_path: Path) -> None:
    """Direct service call: invalid transition raises ValueError."""
    import pytest

    db = tmp_path / "test_invalid.db"
    svc = TaskLifecycleService(db)

    task = svc.create_task(
        "proj-invalid",
        TaskCreate(title="Invalid test", status="not_started"),
    )
    svc.transition_status(
        "proj-invalid",
        task.id,
        TaskStatusTransition(status="cancelled", reason="Done"),
    )

    with pytest.raises(ValueError):
        svc.transition_status(
            "proj-invalid",
            task.id,
            TaskStatusTransition(status="completed", reason="Try reactivate"),
        )


def test_service_not_found_raises(tmp_path: Path) -> None:
    """Getting a non-existent task raises LookupError."""
    import pytest

    db = tmp_path / "test_404.db"
    svc = TaskLifecycleService(db)

    with pytest.raises(LookupError):
        svc.get_task("any-proj", "no-such-task")


def test_parse_csv_util() -> None:
    """_parse_csv helper returns expected rows."""
    data = b"title,owner\nTask X,alice\nTask Y,bob\n"
    rows = _parse_csv(data)
    assert len(rows) == 2
    assert rows[0]["title"] == "Task X"
    assert rows[1]["title"] == "Task Y"


def test_parse_xlsx_util() -> None:
    """_parse_xlsx helper returns expected rows."""
    try:
        from openpyxl import Workbook
    except ImportError:
        import pytest
        pytest.skip("openpyxl not installed")

    wb = Workbook()
    ws = wb.active
    ws.append(["title", "owner"])
    ws.append(["XLSX A", "alice"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    rows = _parse_xlsx(buf.read())
    assert len(rows) == 1
    assert rows[0]["title"] == "XLSX A"


# ============================================================================
# Extraction edge cases
# ============================================================================


def test_extract_empty_text_yields_no_candidates(tmp_path: Path) -> None:
    """Empty text produces zero candidates."""
    app = create_app(_settings(tmp_path))
    with TestClient(app) as client:
        _setup_project(client)
        result = client.post(
            "/api/projects/demo-project/tasks/extract",
            json={"source_text": "", "source_kind": "meeting_notes", "project_id": "demo-project"},
        ).json()
        assert len(result["candidates"]) == 0


def test_extract_no_action_items_yields_no_candidates(tmp_path: Path) -> None:
    """Text with no action items yields zero candidates."""
    app = create_app(_settings(tmp_path))
    with TestClient(app) as client:
        _setup_project(client)
        result = client.post(
            "/api/projects/demo-project/tasks/extract",
            json={
                "source_text": "The weather is nice today. We had a great lunch.",
                "source_kind": "meeting_notes",
                "project_id": "demo-project",
            },
        ).json()
        assert len(result["candidates"]) == 0

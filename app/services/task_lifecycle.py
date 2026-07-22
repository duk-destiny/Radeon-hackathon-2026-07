"""Phase F — Task lifecycle service: CRUD, state machine, confirmation, import.

Responsibilities
----------------
* Create / edit tasks with SQLite persistence
* Task state machine with valid-transition checks
* Candidate task extraction from text (meeting notes, requirements, reports)
* Human confirmation queue (accept / modify / ignore)
* CSV / XLSX import with diff preview, dedup, and confirmation gating
* Operation audit trail
"""

from __future__ import annotations

import csv
import hashlib
import io
import json
import logging
import uuid
from collections.abc import Sequence
from datetime import date, datetime, timezone
from pathlib import Path
from sqlite3 import Connection, IntegrityError

from app.schemas.models import (
    CandidateTask,
    ConfirmationAction,
    ConfirmationRecord,
    OperationAuditRecord,
    TaskChangeRecord,
    TaskCreate,
    TaskExtractionResult,
    TaskImportConfirm,
    TaskImportDiff,
    TaskImportResult,
    TaskRecord,
    TaskStatusTransition,
    TaskUpdate,
)
from app.schemas.task_sql import DDL, is_valid_transition

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# helpers
# ---------------------------------------------------------------------------


def _now_iso() -> str:
    return datetime.now(tz=timezone.utc).isoformat(timespec="seconds")


def _to_json(data: object) -> str:
    return json.dumps(data, ensure_ascii=False, default=str)


def _from_json(raw: str | None, default: object = None) -> object:
    if raw is None:
        return default
    try:
        return json.loads(raw)
    except (json.JSONDecodeError, TypeError):
        return default


def _make_fingerprint(row: dict[str, str]) -> str:
    """Deterministic dedup key for a task row."""
    norm = {k.lower().strip(): (v or "").strip() for k, v in row.items()}
    payload = json.dumps(norm, sort_keys=True, ensure_ascii=False)
    return hashlib.sha256(payload.encode()).hexdigest()


# ---------------------------------------------------------------------------
# SQLite bootstrap
# ---------------------------------------------------------------------------


class TaskLifecycleService:
    """Manages the full task lifecycle backed by the project SQLite DB."""

    def __init__(self, db_path: Path | str) -> None:
        self.db_path = Path(db_path)
        self.db_path.parent.mkdir(parents=True, exist_ok=True)
        self._ensure_schema()

    # -- connection management -------------------------------------------------

    def _connect(self) -> Connection:
        conn = Connection(str(self.db_path))
        conn.execute("PRAGMA journal_mode=WAL")
        conn.execute("PRAGMA foreign_keys=ON")
        conn.row_factory = lambda c, r: r  # plain tuples for simplicity
        return conn

    def _ensure_schema(self) -> None:
        with self._connect() as conn:
            conn.executescript(DDL)
            conn.commit()

    # -- task CRUD ------------------------------------------------------------

    def create_task(self, project_id: str, req: TaskCreate) -> TaskRecord:
        """Create a formal task. If status is *pending_confirmation*, a
        confirmation record is also created."""
        task_id = str(uuid.uuid4())
        now = _now_iso()

        deps_json = _to_json(req.dependencies)
        due_str = req.due_date.isoformat() if req.due_date else None

        with self._connect() as conn:
            conn.execute(
                """INSERT INTO task (id, project_id, title, owner, due_date,
                   priority, acceptance_criteria, dependencies, source_ref,
                   status, created_at, updated_at)
                   VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""",
                (
                    task_id,
                    project_id,
                    req.title,
                    req.owner,
                    due_str,
                    req.priority,
                    req.acceptance_criteria,
                    deps_json,
                    req.source_ref,
                    req.status.value,
                    now,
                    now,
                ),
            )
            # record initial status as event
            conn.execute(
                """INSERT INTO task_change (task_id, project_id,
                   from_status, to_status, changed_by, change_reason, changed_at)
                   VALUES (?,?,?,?,?,?,?)""",
                (task_id, project_id, None, req.status.value, None, "initial creation", now),
            )
            # if created as pending_confirmation, add to confirmation queue
            if req.status.value == "pending_confirmation":
                conn.execute(
                    """INSERT INTO confirmation (task_id, project_id,
                       candidate_title, candidate_owner, candidate_due_date,
                       candidate_priority, candidate_acceptance,
                       candidate_dependencies, source_ref, source_kind,
                       confidence, status, created_at)
                       VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                    (
                        task_id,
                        project_id,
                        req.title,
                        req.owner,
                        due_str,
                        req.priority,
                        req.acceptance_criteria,
                        deps_json,
                        req.source_ref,
                        "manual",
                        1.0,
                        "pending",
                        now,
                    ),
                )
            conn.commit()

        self._audit(project_id, "task", task_id, "create", None)
        return self.get_task(project_id, task_id)

    def get_task(self, project_id: str, task_id: str) -> TaskRecord:
        """Return a single task by id or raise LookupError."""
        with self._connect() as conn:
            row = conn.execute(
                "SELECT * FROM task WHERE id = ? AND project_id = ?",
                (task_id, project_id),
            ).fetchone()
        if row is None:
            raise LookupError(f"Task {task_id} not found in project {project_id}")
        return _task_row_to_record(row)

    def list_tasks(self, project_id: str, status_filter: str | None = None) -> list[TaskRecord]:
        """List tasks for a project, optionally filtered by status."""
        with self._connect() as conn:
            if status_filter:
                rows = conn.execute(
                    "SELECT * FROM task WHERE project_id = ? AND status = ? ORDER BY created_at DESC",
                    (project_id, status_filter),
                ).fetchall()
            else:
                rows = conn.execute(
                    "SELECT * FROM task WHERE project_id = ? ORDER BY created_at DESC",
                    (project_id,),
                ).fetchall()
        return [_task_row_to_record(r) for r in rows]

    def update_task(self, project_id: str, task_id: str, req: TaskUpdate) -> TaskRecord:
        """Edit a task's mutable fields (title, owner, due_date, priority, etc.)."""
        self.get_task(project_id, task_id)  # ensure exists
        now = _now_iso()

        fields: list[str] = []
        values: list[str | None] = []

        if req.title is not None:
            fields.append("title = ?")
            values.append(req.title)
        if req.owner is not None:
            fields.append("owner = ?")
            values.append(req.owner)
        if req.due_date is not None:
            fields.append("due_date = ?")
            values.append(req.due_date.isoformat())
        if req.priority is not None:
            fields.append("priority = ?")
            values.append(req.priority)
        if req.acceptance_criteria is not None:
            fields.append("acceptance_criteria = ?")
            values.append(req.acceptance_criteria)
        if req.dependencies is not None:
            fields.append("dependencies = ?")
            values.append(_to_json(req.dependencies))
        if req.source_ref is not None:
            fields.append("source_ref = ?")
            values.append(req.source_ref)

        if fields:
            fields.append("updated_at = ?")
            values.append(now)
            values.extend([task_id, project_id])
            with self._connect() as conn:
                conn.execute(
                    f"UPDATE task SET {', '.join(fields)} WHERE id = ? AND project_id = ?",
                    values,
                )
                conn.commit()

        self._audit(project_id, "task", task_id, "update", None)
        return self.get_task(project_id, task_id)

    # -- state machine --------------------------------------------------------

    def transition_status(
        self, project_id: str, task_id: str, req: TaskStatusTransition
    ) -> TaskRecord:
        """Move a task from its current status to a new status, validating the
        transition against the allowed state graph."""
        task = self.get_task(project_id, task_id)
        from_status = task.status.value if hasattr(task.status, "value") else task.status

        if not is_valid_transition(from_status, req.status.value):
            raise ValueError(
                f"Invalid transition: {from_status} → {req.status.value}"
            )

        now = _now_iso()
        with self._connect() as conn:
            conn.execute(
                "UPDATE task SET status = ?, updated_at = ? WHERE id = ? AND project_id = ?",
                (req.status.value, now, task_id, project_id),
            )
            conn.execute(
                """INSERT INTO task_change (task_id, project_id,
                   from_status, to_status, changed_by, change_reason, changed_at)
                   VALUES (?,?,?,?,?,?,?)""",
                (task_id, project_id, from_status, req.status.value, req.changed_by, req.reason, now),
            )
            conn.commit()

        self._audit(project_id, "task", task_id, "transition", req.changed_by)
        return self.get_task(project_id, task_id)

    def get_task_history(self, project_id: str, task_id: str) -> list[TaskChangeRecord]:
        """Return all status change events for a task."""
        with self._connect() as conn:
            rows = conn.execute(
                "SELECT * FROM task_change WHERE task_id = ? AND project_id = ? ORDER BY changed_at",
                (task_id, project_id),
            ).fetchall()
        return [_change_row_to_record(r) for r in rows]

    # -- candidate extraction -------------------------------------------------

    def extract_candidates(
        self, project_id: str, source_text: str, source_kind: str = "meeting_notes"
    ) -> TaskExtractionResult:
        """Heuristic extraction of candidate tasks from unstructured text.

        The extraction looks for lines that match task-like patterns
        (action items, TODOs, deadlines). In production this would be
        replaced by an LLM call — the current implementation is a
        rule-based baseline.
        """
        import re

        candidates: list[CandidateTask] = []
        lines = source_text.split("\n")

        # pattern 1: numbered action items
        action_pat = re.compile(
            r"^\s*(?:\d+[.)]\s*|[*-]\s*|(?:(?:TODO|ACTION|TASK)\s*[:：]\s*))(.*)",
            re.IGNORECASE,
        )
        # pattern 2: deadline mentions
        deadline_pat = re.compile(
            r"(?:by|due|deadline|截止|到期)[\s:：]*(\d{4}[-/]\d{1,2}[-/]\d{1,2})",
            re.IGNORECASE,
        )
        # pattern 3: owner mentions
        owner_pat = re.compile(
            r"@(\w[\w.]*)", re.IGNORECASE,
        )

        for line in lines:
            m = action_pat.match(line.strip())
            if not m:
                continue
            title = m.group(1).strip()
            if len(title) < 3:
                continue

            owner_match = owner_pat.search(title)
            owner = owner_match.group(1) if owner_match else None

            deadline_match = deadline_pat.search(title)
            due_date = None
            if deadline_match:
                try:
                    due_date = date.fromisoformat(deadline_match.group(1).replace("/", "-"))
                except ValueError:
                    pass

            # confidence heuristic: higher when both owner and deadline present
            confidence = 0.5
            if owner:
                confidence += 0.2
            if due_date:
                confidence += 0.15
            if len(title) > 20:
                confidence += 0.15

            candidates.append(
                CandidateTask(
                    title=title,
                    owner=owner,
                    due_date=due_date,
                    source_kind=source_kind,
                    confidence=min(confidence, 1.0),
                )
            )

        logger.info(
            "extracted %d candidate tasks from source kind=%s project=%s",
            len(candidates),
            source_kind,
            project_id,
        )
        return TaskExtractionResult(candidates=candidates)

    def submit_candidates(
        self, project_id: str, candidates: list[CandidateTask]
    ) -> list[ConfirmationRecord]:
        """Submit extracted candidates to the confirmation queue.

        Each candidate creates a task with *pending_confirmation* status
        and a confirmation record.
        """
        records: list[ConfirmationRecord] = []
        for candidate in candidates:
            task_id = str(uuid.uuid4())
            now = _now_iso()
            deps_json = _to_json(candidate.dependencies)
            due_str = candidate.due_date.isoformat() if candidate.due_date else None

            with self._connect() as conn:
                conn.execute(
                    """INSERT INTO task (id, project_id, title, owner, due_date,
                       priority, acceptance_criteria, dependencies, source_ref,
                       status, created_at, updated_at)
                       VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""",
                    (
                        task_id,
                        project_id,
                        candidate.title,
                        candidate.owner,
                        due_str,
                        candidate.priority,
                        candidate.acceptance_criteria,
                        deps_json,
                        candidate.source_ref,
                        "pending_confirmation",
                        now,
                        now,
                    ),
                )
                conn.execute(
                    """INSERT INTO task_change
                       (task_id, project_id, from_status, to_status,
                        changed_by, change_reason, changed_at)
                       VALUES (?,?,?,?,?,?,?)""",
                    (task_id, project_id, None, "pending_confirmation", None, "candidate extraction", now),
                )
                conn.execute(
                    """INSERT INTO confirmation
                       (task_id, project_id, candidate_title, candidate_owner,
                        candidate_due_date, candidate_priority,
                        candidate_acceptance, candidate_dependencies,
                        source_ref, source_kind, confidence, status, created_at)
                       VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                    (
                        task_id,
                        project_id,
                        candidate.title,
                        candidate.owner,
                        due_str,
                        candidate.priority,
                        candidate.acceptance_criteria,
                        deps_json,
                        candidate.source_ref,
                        candidate.source_kind,
                        candidate.confidence,
                        "pending",
                        now,
                    ),
                )
                conn.commit()

            records.append(self._get_confirmation_record(project_id, task_id))
            self._audit(project_id, "candidate", task_id, "submit", None)

        return records

    # -- confirmation queue ---------------------------------------------------

    def list_confirmation_queue(
        self, project_id: str, status_filter: str | None = None
    ) -> list[ConfirmationRecord]:
        """List items in the human confirmation queue."""
        with self._connect() as conn:
            if status_filter:
                rows = conn.execute(
                    "SELECT * FROM confirmation WHERE project_id = ? AND status = ? ORDER BY created_at DESC",
                    (project_id, status_filter),
                ).fetchall()
            else:
                rows = conn.execute(
                    "SELECT * FROM confirmation WHERE project_id = ? ORDER BY created_at DESC",
                    (project_id,),
                ).fetchall()
        return [_conf_row_to_record(r) for r in rows]

    def process_confirmation(
        self, project_id: str, task_id: str, req: ConfirmationAction
    ) -> ConfirmationRecord:
        """Process a human confirmation action (accept / modify / ignore)."""
        # verify confirmation exists and is pending
        with self._connect() as conn:
            row = conn.execute(
                "SELECT * FROM confirmation WHERE task_id = ? AND project_id = ?",
                (task_id, project_id),
            ).fetchone()
            if row is None:
                raise LookupError(f"Confirmation for task {task_id} not found")
            rec = _conf_row_to_record(row)
            if rec.status != "pending":
                raise ValueError(f"Confirmation already processed: {rec.status}")

        now = _now_iso()

        if req.action == "ignore":
            # cancel the task, mark confirmation as ignored
            with self._connect() as conn:
                conn.execute(
                    "UPDATE confirmation SET status = 'ignored', confirmed_by = ?, "
                    "confirmation_basis = ?, confirmation_notes = ?, confirmed_at = ? WHERE task_id = ? AND project_id = ?",
                    (req.confirmed_by, req.confirmation_basis, req.confirmation_notes, now, task_id, project_id),
                )
                conn.execute(
                    "UPDATE task SET status = 'cancelled', updated_at = ? WHERE id = ? AND project_id = ?",
                    (now, task_id, project_id),
                )
                conn.execute(
                    """INSERT INTO task_change (task_id, project_id, from_status, to_status,
                       changed_by, change_reason, changed_at)
                       VALUES (?,?,?,?,?,?,?)""",
                    (task_id, project_id, "pending_confirmation", "cancelled", req.confirmed_by, "confirmation ignored", now),
                )
                conn.commit()
            self._audit(project_id, "confirmation", task_id, "ignore", req.confirmed_by)
            return self._get_confirmation_record(project_id, task_id)

        elif req.action == "accept":
            # move task to not_started
            with self._connect() as conn:
                conn.execute(
                    "UPDATE confirmation SET status = 'accepted', confirmed_by = ?, "
                    "confirmation_basis = ?, confirmation_notes = ?, confirmed_at = ? WHERE task_id = ? AND project_id = ?",
                    (req.confirmed_by, req.confirmation_basis, req.confirmation_notes, now, task_id, project_id),
                )
                conn.execute(
                    """UPDATE task SET status = 'not_started', confirmed_by = ?,
                       confirmed_at = ?, confirmation_basis = ?,
                       confirmation_notes = ?, updated_at = ?
                       WHERE id = ? AND project_id = ?""",
                    (req.confirmed_by, now, req.confirmation_basis, req.confirmation_notes, now, task_id, project_id),
                )
                conn.execute(
                    """INSERT INTO task_change (task_id, project_id, from_status, to_status,
                       changed_by, change_reason, changed_at)
                       VALUES (?,?,?,?,?,?,?)""",
                    (task_id, project_id, "pending_confirmation", "not_started", req.confirmed_by, "confirmation accepted", now),
                )
                conn.commit()
            self._audit(project_id, "confirmation", task_id, "accept", req.confirmed_by)
            return self._get_confirmation_record(project_id, task_id)

        else:  # modify
            # apply modifications then accept
            with self._connect() as conn:
                # apply modified fields to task
                if req.modified_title is not None:
                    conn.execute("UPDATE task SET title = ? WHERE id = ? AND project_id = ?", (req.modified_title, task_id, project_id))
                if req.modified_owner is not None:
                    conn.execute("UPDATE task SET owner = ? WHERE id = ? AND project_id = ?", (req.modified_owner, task_id, project_id))
                if req.modified_due_date is not None:
                    conn.execute("UPDATE task SET due_date = ? WHERE id = ? AND project_id = ?", (req.modified_due_date.isoformat(), task_id, project_id))
                if req.modified_priority is not None:
                    conn.execute("UPDATE task SET priority = ? WHERE id = ? AND project_id = ?", (req.modified_priority, task_id, project_id))
                if req.modified_acceptance is not None:
                    conn.execute("UPDATE task SET acceptance_criteria = ? WHERE id = ? AND project_id = ?", (req.modified_acceptance, task_id, project_id))
                if req.modified_dependencies is not None:
                    conn.execute("UPDATE task SET dependencies = ? WHERE id = ? AND project_id = ?", (_to_json(req.modified_dependencies), task_id, project_id))

                # mark confirmation accepted
                conn.execute(
                    "UPDATE confirmation SET status = 'accepted', confirmed_by = ?, "
                    "confirmation_basis = ?, confirmation_notes = ?, confirmed_at = ? WHERE task_id = ? AND project_id = ?",
                    (req.confirmed_by, req.confirmation_basis, req.confirmation_notes, now, task_id, project_id),
                )
                conn.execute(
                    """UPDATE task SET status = 'not_started', confirmed_by = ?,
                       confirmed_at = ?, confirmation_basis = ?,
                       confirmation_notes = ?, updated_at = ?
                       WHERE id = ? AND project_id = ?""",
                    (req.confirmed_by, now, req.confirmation_basis, req.confirmation_notes, now, task_id, project_id),
                )
                conn.execute(
                    """INSERT INTO task_change (task_id, project_id, from_status, to_status,
                       changed_by, change_reason, changed_at)
                       VALUES (?,?,?,?,?,?,?)""",
                    (task_id, project_id, "pending_confirmation", "not_started", req.confirmed_by, "confirmation accepted with modifications", now),
                )
                conn.commit()
            self._audit(project_id, "confirmation", task_id, "modify_accept", req.confirmed_by)
            return self._get_confirmation_record(project_id, task_id)

    # -- import (CSV / XLSX) --------------------------------------------------

    def preview_import(
        self, project_id: str, file_bytes: bytes, filename: str
    ) -> tuple[TaskImportDiff, list[CandidateTask]]:
        """Parse a CSV or XLSX file and return a diff preview plus candidate
        tasks. Does NOT persist anything."""
        ext = Path(filename).suffix.lower()
        rows: list[dict[str, str]] = []

        if ext == ".csv":
            rows = _parse_csv(file_bytes)
        elif ext in (".xlsx", ".xls"):
            rows = _parse_xlsx(file_bytes)
        else:
            raise ValueError(f"Unsupported format: {ext}")

        candidates: list[CandidateTask] = []
        new_count = 0
        dup_count = 0
        conflict_count = 0
        preview_rows: list[dict[str, str]] = []

        # collect existing fingerprints for dedup
        existing_fps: set[str] = set()
        with self._connect() as conn:
            fp_rows = conn.execute(
                "SELECT fingerprint FROM task_import_fingerprint WHERE project_id = ?",
                (project_id,),
            ).fetchall()
            existing_fps = {r[0] for r in fp_rows}

            existing_titles: set[str] = set()
            title_rows = conn.execute(
                "SELECT LOWER(title) FROM task WHERE project_id = ?", (project_id,)
            ).fetchall()
            existing_titles = {r[0].lower().strip() for r in title_rows}

        for row in rows:
            fp = _make_fingerprint(row)
            title = (row.get("title") or row.get("任务名称") or row.get("name") or "").strip()

            if fp in existing_fps:
                dup_count += 1
                continue

            if title.lower() in existing_titles:
                conflict_count += 1
                preview_rows.append(row)
                continue

            new_count += 1
            candidate = CandidateTask(
                title=title,
                owner=row.get("owner") or row.get("负责人") or row.get("assignee"),
                due_date=_parse_date(row.get("due_date") or row.get("截止日期") or row.get("deadline")),
                priority=(row.get("priority") or row.get("优先级") or "").strip() or None,
                acceptance_criteria=row.get("acceptance_criteria") or row.get("验收标准"),
                source_ref=f"import:{filename}",
                source_kind="file_import",
                confidence=0.9,
            )
            candidates.append(candidate)
            preview_rows.append(row)

        diff = TaskImportDiff(
            new_rows=new_count,
            duplicate_rows=dup_count,
            conflict_rows=conflict_count,
            preview=preview_rows[:20],  # limit preview
        )
        return diff, candidates

    def confirm_import(
        self,
        project_id: str,
        candidates: list[CandidateTask],
        file_bytes: bytes,
        filename: str,
        req: TaskImportConfirm,
    ) -> TaskImportResult:
        """Confirm and execute a CSV/XLSX import. Records fingerprints for
        future dedup.

        The function parses the file independently so that skipped/
        duplicate rows are properly counted even when the preview has
        already filtered them out.
        """
        imported = 0
        skipped = 0
        errors = 0
        details: list[str] = []

        ext = Path(filename).suffix.lower()
        if ext == ".csv":
            rows = _parse_csv(file_bytes)
        elif ext in (".xlsx", ".xls"):
            rows = _parse_xlsx(file_bytes)
        else:
            raise ValueError(f"Unsupported format: {ext}")

        # build a set of candidate titles from the preview (for dedup reference)
        candidate_titles = {c.title.lower().strip() for c in candidates}
        # also build candidate fingerprints for quick lookup
        candidate_fps: dict[str, CandidateTask] = {}
        for c in candidates:
            fp = _make_fingerprint(
                {
                    "title": c.title,
                    "owner": c.owner or "",
                    "due_date": c.due_date.isoformat() if c.due_date else "",
                    "priority": c.priority or "",
                }
            )
            candidate_fps[fp] = c

        for idx, row in enumerate(rows):
            title = (row.get("title") or row.get("任务名称") or row.get("name") or "").strip()
            if not title:
                skipped += 1
                details.append(f"Row {idx}: empty title — skipped")
                continue

            fp = _make_fingerprint(row)

            if req.skip_duplicates:
                with self._connect() as conn:
                    existing = conn.execute(
                        "SELECT id FROM task_import_fingerprint WHERE project_id = ? AND fingerprint = ?",
                        (project_id, fp),
                    ).fetchone()
                    if existing:
                        skipped += 1
                        details.append(f"Row {idx}: duplicate skipped (fingerprint match)")
                        continue

            try:
                owner = (row.get("owner") or row.get("负责人") or row.get("assignee") or "").strip() or None
                due = _parse_date(row.get("due_date") or row.get("截止日期") or row.get("deadline"))
                priority = (row.get("priority") or row.get("优先级") or "").strip() or None
                acceptance = (row.get("acceptance_criteria") or row.get("验收标准") or "").strip() or None

                if req.overwrite_conflicts:
                    # check if a task with same title exists
                    with self._connect() as conn:
                        existing_task = conn.execute(
                            "SELECT id FROM task WHERE LOWER(title) = ? AND project_id = ?",
                            (title.lower(), project_id),
                        ).fetchone()
                        if existing_task:
                            conn.execute(
                                "UPDATE task SET owner = ?, due_date = ?, priority = ?, "
                                "acceptance_criteria = ?, updated_at = ? WHERE id = ?",
                                (owner, due.isoformat() if due else None, priority,
                                 acceptance, _now_iso(), existing_task[0]),
                            )
                            conn.commit()
                            imported += 1
                            details.append(f"Row {idx}: overwritten existing task {existing_task[0]}")
                            continue

                task = self.create_task(
                    project_id,
                    TaskCreate(
                        title=title,
                        owner=owner,
                        due_date=due,
                        priority=priority,
                        acceptance_criteria=acceptance,
                        source_ref=f"import:{filename}",
                        status="not_started",  # imported tasks skip confirmation
                    ),
                )

                # record fingerprint
                with self._connect() as conn:
                    try:
                        conn.execute(
                            "INSERT INTO task_import_fingerprint (project_id, fingerprint, task_id) VALUES (?,?,?)",
                            (project_id, fp, task.id),
                        )
                        conn.commit()
                    except IntegrityError:
                        pass  # fingerprint already exists (race condition)

                imported += 1
                details.append(f"Row {idx}: imported as {task.id}")

            except Exception as e:
                errors += 1
                details.append(f"Row {idx}: error - {e}")

        self._audit(
            project_id,
            "import",
            filename,
            "confirm_import",
            req.confirmed_by,
        )
        return TaskImportResult(imported=imported, skipped=skipped, errors=errors, details=details)

    # -- audit -----------------------------------------------------------------

    def _audit(
        self,
        project_id: str,
        entity_type: str,
        entity_id: str,
        operation: str,
        operator: str | None,
    ) -> None:
        now = _now_iso()
        with self._connect() as conn:
            conn.execute(
                """INSERT INTO operation_audit
                   (project_id, entity_type, entity_id, operation, operator, created_at)
                   VALUES (?,?,?,?,?,?)""",
                (project_id, entity_type, entity_id, operation, operator, now),
            )
            conn.commit()

    def get_audit_log(self, project_id: str, limit: int = 100) -> list[OperationAuditRecord]:
        """Return recent audit entries for a project."""
        with self._connect() as conn:
            rows = conn.execute(
                "SELECT * FROM operation_audit WHERE project_id = ? ORDER BY created_at DESC LIMIT ?",
                (project_id, limit),
            ).fetchall()
        return [_audit_row_to_record(r) for r in rows]

    # -- internal helpers -----------------------------------------------------

    def _get_confirmation_record(self, project_id: str, task_id: str) -> ConfirmationRecord:
        with self._connect() as conn:
            row = conn.execute(
                "SELECT * FROM confirmation WHERE task_id = ? AND project_id = ?",
                (task_id, project_id),
            ).fetchone()
            if row is None:
                raise LookupError(f"Confirmation for task {task_id} not found")
            return _conf_row_to_record(row)


# ---------------------------------------------------------------------------
# row → record mappers
# ---------------------------------------------------------------------------


def _task_row_to_record(row: tuple) -> TaskRecord:
    """Map a task SQLite row to a TaskRecord."""
    cols = [d[0] for d in row if hasattr(d, "__getitem__")]
    if not cols:
        return TaskRecord(
            id=str(row[0]),
            project_id=str(row[1]),
            title=str(row[2]),
            owner=row[3] if row[3] else None,
            due_date=row[4] if row[4] else None,
            priority=row[5] if row[5] else None,
            acceptance_criteria=row[6] if row[6] else None,
            dependencies=list(_from_json(row[7], [])),
            source_ref=row[8] if row[8] else None,
            status=row[9] if row[9] else "pending_confirmation",
            confirmed_by=row[10] if row[10] else None,
            confirmed_at=row[11] if row[11] else None,
            confirmation_basis=row[12] if row[12] else None,
            confirmation_notes=row[13] if row[13] else None,
            created_at=str(row[14]) if row[14] else "",
            updated_at=str(row[15]) if row[15] else "",
        )
    return TaskRecord(
        id=str(row[0]),
        project_id=str(row[1]),
        title=str(row[2]),
        owner=str(row[3]) if row[3] else None,
        due_date=str(row[4]) if row[4] else None,
        priority=str(row[5]) if row[5] else None,
        acceptance_criteria=str(row[6]) if row[6] else None,
        dependencies=list(_from_json(str(row[7]), [])),
        source_ref=str(row[8]) if row[8] else None,
        status=str(row[9]) if row[9] else "pending_confirmation",
        confirmed_by=str(row[10]) if row[10] else None,
        confirmed_at=str(row[11]) if row[11] else None,
        confirmation_basis=str(row[12]) if row[12] else None,
        confirmation_notes=str(row[13]) if row[13] else None,
        created_at=str(row[14]) if row[14] else "",
        updated_at=str(row[15]) if row[15] else "",
    )


def _conf_row_to_record(row: tuple) -> ConfirmationRecord:
    """Map a confirmation SQLite row to a ConfirmationRecord."""
    return ConfirmationRecord(
        id=int(row[0]),
        task_id=str(row[1]),
        project_id=str(row[2]),
        candidate_title=str(row[3]),
        candidate_owner=str(row[4]) if row[4] else None,
        candidate_due_date=str(row[5]) if row[5] else None,
        candidate_priority=str(row[6]) if row[6] else None,
        candidate_acceptance=str(row[7]) if row[7] else None,
        candidate_dependencies=list(_from_json(str(row[8]), [])),
        source_ref=str(row[9]) if row[9] else None,
        source_kind=str(row[10]),
        confidence=float(row[11]) if row[11] else 0.5,
        status=str(row[12]),
        confirmed_by=str(row[13]) if row[13] else None,
        confirmation_basis=str(row[14]) if row[14] else None,
        confirmation_notes=str(row[15]) if row[15] else None,
        confirmed_at=str(row[16]) if row[16] else None,
        created_at=str(row[17]) if row[17] else "",
    )


def _change_row_to_record(row: tuple) -> TaskChangeRecord:
    """Map a task_change SQLite row to a TaskChangeRecord."""
    return TaskChangeRecord(
        id=int(row[0]),
        task_id=str(row[1]),
        project_id=str(row[2]),
        from_status=str(row[3]) if row[3] else None,
        to_status=str(row[4]),
        changed_by=str(row[5]) if row[5] else None,
        change_reason=str(row[6]) if row[6] else None,
        changed_at=str(row[7]) if row[7] else "",
    )


def _audit_row_to_record(row: tuple) -> OperationAuditRecord:
    """Map an operation_audit SQLite row to an OperationAuditRecord."""
    return OperationAuditRecord(
        id=int(row[0]),
        project_id=str(row[1]),
        entity_type=str(row[2]),
        entity_id=str(row[3]),
        operation=str(row[4]),
        operator=str(row[5]) if row[5] else None,
        details=str(row[6]) if row[6] else None,
        created_at=str(row[7]) if row[7] else "",
    )


# ---------------------------------------------------------------------------
# file-parsing helpers
# ---------------------------------------------------------------------------


def _parse_csv(file_bytes: bytes) -> list[dict[str, str]]:
    """Parse CSV bytes into a list of row dicts."""
    text = file_bytes.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    return [dict(row) for row in reader]


def _parse_xlsx(file_bytes: bytes) -> list[dict[str, str]]:
    """Parse XLSX bytes into a list of row dicts."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ImportError(
            "openpyxl is required for XLSX import. Install it with: pip install openpyxl"
        ) from exc

    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(next(rows_iter))]
    result: list[dict[str, str]] = []
    for row in rows_iter:
        result.append({headers[i]: (str(v) if v is not None else "") for i, v in enumerate(row)})
    wb.close()
    return result


def _parse_date(raw: str | None) -> date | None:
    """Try to parse a date from string."""
    if not raw:
        return None
    raw = raw.strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%m/%d/%Y", "%d/%m/%Y", "%Y%m%d"):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            continue
    # fallback: isoformat-like
    try:
        return date.fromisoformat(raw)
    except ValueError:
        pass
    return None
